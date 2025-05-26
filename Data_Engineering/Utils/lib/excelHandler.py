import os
from traceback import print_exc
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelHandler:
    def __init__(self):
        self.workbook = None
        self.file_path = None

    def reset(self):
        if self.workbook:
            self.workbook.close()
            print(f"Workbook {self.file_path} closed.")
        self.workbook = None
        self.file_path = None

    def load_workbook(self, file_path: str, keep_vba: bool = True, keep_links: bool = False) -> bool:
        if os.path.exists(file_path):
            self.workbook = load_workbook(file_path, keep_links=keep_links, keep_vba=keep_vba)
            self.file_path = file_path
            return True
        print(f"File {file_path} does not exist.")
        return False

    def create_workbook(self, file_path: str, keep_vba: bool = True) -> bool:
        if os.path.exists(file_path):
            print(f"File {file_path} already exists.")
            return False
        self.workbook = Workbook()
        self.file_path = file_path
        if keep_vba:
            self.workbook.save(file_path)
        return True

    def save_workbook(self, file_path: str = None) -> bool:
        try:
            target_path = file_path or self.file_path
            if self.workbook and target_path:
                self.workbook.save(target_path)
                return True
            print("No valid file path or workbook to save.")
        except Exception as e:
            print(f"Error saving workbook: {e}")
            print_exc()
        return False

    def get_sheet_names(self) -> list:
        return self.workbook.sheetnames if self.workbook else []

    def get_sheet(self, sheet_name: str):
        if self.workbook and sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        print(f"Sheet {sheet_name} does not exist.")
        return None

    def create_sheet(self, sheet_name: str) -> bool:
        if not self.workbook:
            print("No workbook loaded.")
            return False
        if sheet_name in self.workbook.sheetnames:
            print(f"Sheet {sheet_name} already exists.")
            return False
        self.workbook.create_sheet(sheet_name)
        return True

    def delete_sheet(self, sheet_name: str) -> bool:
        try:
            sheet = self.get_sheet(sheet_name)
            if sheet:
                self.workbook.remove(sheet)
                return True
            return False
        except Exception as e:
            print(f"Error deleting sheet: {e}")
            print_exc()
            return False

    def reset_sheet(self, sheet_name: str) -> bool:
        return self.delete_sheet(sheet_name) and self.create_sheet(sheet_name)

    def write_df_to_sheet(self, sheet_name: str, data: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> bool:
        try:
            if not self.workbook:
                print("No workbook loaded.")
                return False

            if sheet_name not in self.workbook.sheetnames:
                if not self.create_sheet(sheet_name):
                    print(f"Failed to create sheet {sheet_name}.")
                    return False

            if not self.reset_sheet(sheet_name):
                print(f"Failed to reset sheet {sheet_name}.")
                return False

            # Convert timezone-aware datetime columns to naive
            for col in data.select_dtypes(include=['datetime64[ns, UTC]', 'datetime64[ns]']).columns:
                if str(data[col].dtype) == "datetime64[ns, UTC]":
                    data[col] = data[col].dt.tz_convert(None)
                else:
                    data[col] = data[col].dt.tz_localize(None)

            sheet = self.get_sheet(sheet_name)

            sheet.append(data.columns.tolist())  
            for row in dataframe_to_rows(data, index=False, header=True):
                sheet.append(row)

            return True
        except Exception as e:
            print(f"Error writing DataFrame to sheet: {e}")
            print_exc()
            return False

    def write_value(self, cell: str, value, sheet_name: str = None) -> bool:
        try:
            if not self.workbook:
                print("No workbook loaded.")
                return False
            sheet = self.get_sheet(sheet_name) if sheet_name else self.workbook.active
            if not sheet:
                print(f"Sheet {sheet_name} does not exist.")
                return False
            sheet[cell] = value
            return True
        except Exception as e:
            print(f"Error writing value to cell {cell}: {e}")
            print_exc()
            return False
