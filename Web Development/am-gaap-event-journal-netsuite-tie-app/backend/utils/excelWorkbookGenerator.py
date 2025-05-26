import sys
import os,sys
from io import BytesIO
scriptPath = os.path.realpath(os.path.dirname(sys.argv[0]))
superPath = os.path.realpath(os.path.dirname(scriptPath))
sys.path.append(superPath)
import json
import yaml
import pandas as pd
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
from utils.gcpHandler import GCPHandler
# from utils.excelMacroGenerator import ExcelMacroGenerator
import time
import glob

class ExcelWorkbookGenerator():

    def __init__(self):
        self.GCPHandler = self.getGCPCHandler()
        self.projectId = 'am-finance-forecast'
        self.eventDataset = 'Marketplace_GAAP_Events'
        self.journalDataset = 'Marketplace_GAAP_Journals'
        self.journalEventMapping = self.getMapping()
        # self.excelMacroGenerator = ExcelMacroGenerator()

    def getGCPCHandler(self):
        '''Get GCP credentials from file'''
        try:
            with open('var/bigquery_service_account.json', 'r') as file:
                return GCPHandler(serviceAccountJson= json.load(file))
        except Exception as e:
            try:
                from google.auth import compute_engine
                from google.cloud import bigquery
                credentials = compute_engine.Credentials()
                return GCPHandler(client= bigquery.Client(credentials=credentials))
            except Exception as e:
                print(traceback.format_exc())
                raise Exception(f"Error getting GCP credentials: {e}")

    def getMapping(self) -> dict:
        '''Get journal event mapping from file or API'''
        try:
            journalEventMapping = open("utils/journal_event_mapping.json", 'r')
            journalEventMapping = json.load(journalEventMapping)
        except:
            print("Error getting journal event mapping from file")
            print(traceback.format_exc())
            journalEventMapping = self.getJournalEventMapping()
        return journalEventMapping
    
    def getSheetName(self,tableName : str) -> str:
        '''Get Excel sheet name for a table'''
        tableName = tableName.replace('a_la_carte_and_vip_', '').replace('try_at_home_', '').replace('payment_providers_', '')
        return tableName[:31]

    def getEvents(self,journal : str) -> list[str]:
        '''Get events for a journal'''
        return self.journalEventMapping[journal]['events']

    def getInsertDateFieldName(self, tableId : str) -> str:
        '''Get insert date field name for a table'''
        print(f"Getting insert date field name for {tableId}")
        table_schema = self.GCPHandler.retrieveTableSchema(tableId)
        for field in table_schema:
            if ('insert' in field.lower() or 'je_id' in field.lower()) and table_schema[field] == 'TIMESTAMP':
                return field

    def getTableData(self,tableName : str, tableType : str, date_from : str, date_to : str, bigQueryJobs : dict) -> pd.DataFrame:
        '''Get BigQuery event table'''
        try:
            dataset = self.eventDataset if tableType == 'event' else self.journalDataset
            tableId = f"{self.projectId}.{dataset}.{tableName}"
            if tableType == 'event':
                insertDateFieldName = self.getInsertDateFieldName(tableId)
                query = f"SELECT * FROM `{tableId}` WHERE DATE({insertDateFieldName},'America/New_York') >= '{date_from}' AND DATE({insertDateFieldName},'America/New_York') <= '{date_to}'"
            else:
                query = f"SELECT * FROM `{tableId}` WHERE DATE(Journal_Insert_Timestamp,'America/New_York') >= '{date_from}' AND DATE(Journal_Insert_Timestamp,'America/New_York') <= '{date_to}'"
            query_result, jobId = self.GCPHandler.runQuery(query)
            bigQueryJobs[jobId] = tableName 
            return pd.DataFrame(query_result)
        except Exception as e:
            print(traceback.format_exc())
            raise Exception(f"Error getting table data: {e}")
        
    def getJournalEventMapping(self):
        config_files = glob.glob("utils/journalsConfig/*.yaml")
        config_files = [path.replace("\\", "/") for path in config_files]
        journals = {}
        for config_file in config_files:
            with open(config_file, 'r') as file:
                config = yaml.safe_load(file)
                if 'am-gaap-reporting-dev' in config:
                    key = list(config['am-gaap-reporting-dev'].keys())[0]
                    events = config['am-gaap-reporting-dev'][key].keys()
                    for event in events:
                        journal_table = config['am-gaap-reporting-dev'][key][event]['output_variables']['destination_table_name'] + '_Journal'
                        event_name = list(config['am-gaap-reporting-dev'][key][event]['events'].keys())[0]
                        event_table = config['am-gaap-reporting-dev'][key][event]['events'][event_name]['input_names']['source_table_name']
                        journals[journal_table] = {'events' : 
                            journals.get(journal_table, {}).get('events', []) + [{'event_name': event, 'sub_event_name': event_name, 'event_table': event_table}],
                                                    'config_file': config_file,
                                                    'config':key}
        with open("utils/journal_event_mapping.json", 'w') as file:
            json.dump(journals, file)

        return journals

    def getAccountsCalculation(self,journalTable : str) -> dict:
        '''Get accounts calculation for a journal'''
        try:
            calc = {}
            configFile = self.journalEventMapping[journalTable]['config_file']
            config_name = self.journalEventMapping[journalTable]['config']
            with open(configFile, 'r') as file:
                config = yaml.safe_load(file)
                events = self.getEvents(journalTable)
                for event in events:
                    eventTable = event['event_table']
                    eventName = event['event_name']
                    subEventName = event['sub_event_name']
                    transactionTypesConfig = config['am-gaap-reporting-dev'][config_name][eventName]['events'][subEventName]['Transaction_Type']
                    transactionTypeColumnName = transactionTypesConfig['column_name']

                    transactionTypes = transactionTypesConfig['types']
                    for transactionType in transactionTypes:
                        transactionTypeName = list(transactionType.keys())[0]
                        transactionTypeFields = transactionType[transactionTypeName]['aggregated_columns']  

                        for transactionTypeField in transactionTypeFields:
                            transactionTypeFieldName = list(transactionTypeField.keys())[0]
                            transactionTypeFieldConfig = transactionTypeField[transactionTypeFieldName] 
                            transactionTypeFieldDebitAccount = transactionTypeFieldConfig['debit_credit_accounts'].get('debit', None)
                            transactionTypeCFieldCreditAccount = transactionTypeFieldConfig['debit_credit_accounts'].get('credit', None)


                            if transactionTypeFieldDebitAccount:
                                calc[transactionTypeFieldDebitAccount] = {'debit': calc.get(transactionTypeFieldDebitAccount, {}).get('debit', []) + [{'event_table': eventTable, 'transaction_type': transactionTypeName, 'transaction_type_column_name': transactionTypeColumnName, 'transaction_type_field_name': transactionTypeFieldName}], 
                                                                        'credit': calc.get(transactionTypeFieldDebitAccount, {}).get('credit', []) }

                            if transactionTypeCFieldCreditAccount:
                                calc[transactionTypeCFieldCreditAccount] = {'credit': calc.get(transactionTypeCFieldCreditAccount, {}).get('credit', []) + [{'event_table': eventTable, 'transaction_type': transactionTypeName, 'transaction_type_column_name': transactionTypeColumnName, 'transaction_type_field_name': transactionTypeFieldName}], 
                                                                            'debit': calc.get(transactionTypeCFieldCreditAccount, {}).get('debit', []) }

            return [{'accountName': account, 'debit': calc[account]['debit'], 'credit': calc[account]['credit']} for account in calc]
        except Exception as e:
            print(traceback.format_exc())
            raise Exception(f"Error getting accounts calculation: {e}")

    def writeWorksheet(self, wb: Workbook, sheetName: str, data: pd.DataFrame):
        '''Write DataFrame to an Excel worksheet using openpyxl'''
        try:
            print(f"Writing data to {sheetName}")
            
            # Remove existing sheet if it exists
            if sheetName in wb.sheetnames:
                wb.remove(wb[sheetName])

            # Create new sheet
            ws = wb.create_sheet(title=sheetName)

            # Convert all timezone-aware datetime columns to timezone-naive
            for col in data.select_dtypes(include=['datetime64[ns, UTC]', 'datetime64[ns]']).columns:
                if data[col].dtype.name == "datetime64[ns, UTC]":  # If column has timezone
                    data[col] = data[col].dt.tz_convert(None)  # Convert to naive datetime
                else:
                    data[col] = data[col].dt.tz_localize(None)  # Remove timezone if localized
            
            # Append data to worksheet
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)

            # Auto-adjust column widths
            for col_idx, col in enumerate(data.columns, 1):
                max_length = max(
                    data[col].astype(str).apply(len).max(),  # Max length of column values
                    len(str(col))  # Length of column name
                )
                adjusted_width = max_length + 2  # Add some padding
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Apply formatting to header row
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="156082", end_color="156082", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

        except Exception as e:
            print(traceback.format_exc())
            raise Exception(f"Error writing data to worksheet: {e}")

    def getEventSummaryFormula(self, debitCreditEvent : dict ,eventsData : dict):
        '''Get Excel formula for event summary'''

        if len(debitCreditEvent) == 0:
            return "0"
        
        debitCreditFormula = ""
        for j, debitEvent in enumerate(debitCreditEvent):
            eventTable = debitEvent['event_table']
            dataSourceSheet = self.getSheetName(eventTable)
            transactionType = debitEvent['transaction_type']
            transactionTypeFieldName = debitEvent['transaction_type_field_name']
            transactionTypeColumnName = debitEvent['transaction_type_column_name']


            maxRow = eventsData[eventTable]['length'] + 1 
            transactionTypeColumnIndex = eventsData[eventTable]['data'].columns.get_loc(transactionTypeColumnName) + 1
            transactionTypeFieldIndex = eventsData[eventTable]['data'].columns.get_loc(transactionTypeFieldName) + 1

            debitCreditEventFormula = f"SUMIFS('{dataSourceSheet}'!{get_column_letter(transactionTypeFieldIndex)}2:{get_column_letter(transactionTypeFieldIndex)}{maxRow}, '{dataSourceSheet}'!{get_column_letter(transactionTypeColumnIndex)}2:{get_column_letter(transactionTypeColumnIndex)}{maxRow}, \"{transactionType}\")"

            if j > 0:
                debitCreditFormula += " + "
            debitCreditFormula += debitCreditEventFormula

        return debitCreditFormula

    def getJournalSummaryFormula(self, journal : str, journalData : dict, accountName : str, debitCredit : str):
        '''Get Excel formula for journal summary'''
        journalSheet = self.getSheetName(journal)
        accountNameColumnIndex = journalData['data'].columns.get_loc('Account_Name') + 1
        debitCreditColumnIndex = journalData['data'].columns.get_loc('Debit_Credit') + 1
        valueColumnIndex = journalData['data'].columns.get_loc('Value') + 1

        maxRow = journalData['length'] + 1
        journalSummaryFormula = f"SUMIFS('{journalSheet}'!{get_column_letter(valueColumnIndex)}2:{get_column_letter(valueColumnIndex)}{maxRow}, '{journalSheet}'!{get_column_letter(accountNameColumnIndex)}2:{get_column_letter(accountNameColumnIndex)}{maxRow}, \"{accountName}\", '{journalSheet}'!{get_column_letter(debitCreditColumnIndex)}2:{get_column_letter(debitCreditColumnIndex)}{maxRow}, \"{debitCredit}\")"
        return journalSummaryFormula


    def createSummarySheet(self, workbook: Workbook, journal: str, journalData : pd.DataFrame, eventsData: dict, accountCalculations: list):
        try:
            print("Creating summary sheet")
            summarySheet = workbook.create_sheet(title='Summary', index=0)

            # Headers
            headers = [
                "Account", "Credit - Events", "Credit - Journal", "Credit - NetSuite",
                "Debit - Events", "Debit - Journal", "Debit - NetSuite"
            ]
            
            # Write headers in the first row
            for col_idx, header in enumerate(headers, 1):
                cell = summarySheet.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="156082", end_color="156082", fill_type="solid")  # Background color
                cell.font = Font(color="FFFFFF", bold=True)
            
            nbAccounts = len(accountCalculations)

            for i in range(nbAccounts):
                accountCalculation = accountCalculations[i]
                accountName = accountCalculation['accountName']
                debitEvents = accountCalculation['debit']
                creditEvents = accountCalculation['credit']

                # print(i, accountName)
                # Write account name
                summarySheet.cell(row=i + 2, column=1, value=accountName)

                # Write Credit - Events
                summarySheet.cell(row=i + 2, column=2, value= "= 0 +" + self.getEventSummaryFormula(creditEvents,eventsData))
             
                # Write Credit - Journal
                summarySheet.cell(row=i + 2, column=3, value= "= 0 +" + self.getJournalSummaryFormula(journal, journalData, accountName, "Credit"))

                # Write Credit - NetSuite
                # summarySheet.cell(row=i + 2, column=4, value= "= 0 +" + self.getNetSuiteSummaryFormula(journal, journalData, accountName, "Credit"))

                # Write Debit - Events
                summarySheet.cell(row=i + 2, column=5, value= "= 0 +" + self.getEventSummaryFormula(debitEvents,eventsData))
                
                # Write Debit - Journal
                summarySheet.cell(row=i + 2, column=6, value= "= 0 +" + self.getJournalSummaryFormula(journal, journalData, accountName, "Debit"))

                # Write Debit - NetSuite
                # summarySheet.cell(row=i + 2, column=7, value= "= 0 +" + self.getNetSuiteSummaryFormula(journal, journalData, accountName, "Debit"))

                


            # Write totals
            total_row = nbAccounts + 2
            summarySheet.cell(row=total_row, column=1, value="Total")
            for col in range(2, 8):
                summarySheet.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{nbAccounts+1})")

            # Format
            for row in summarySheet.iter_rows(min_row=1, max_row=nbAccounts + 2, min_col=1, max_col=7):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            # Format to $000,000.00
            for col in range(2, 8):
                for row in range(1, nbAccounts + 3):
                    summarySheet.cell(row=row, column=col).number_format = '$#,##0.00'

            # # Auto-fit columns
            for col in summarySheet.columns:
                # max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = 20
                summarySheet.column_dimensions[col[0].column_letter].width = adjusted_width

            # Add conditional formatting to check if Events = Journal = NetSuite
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

            # Define the last row
            last_row = nbAccounts + 2

            # Conditional Formatting for Credit Columns (B, C, D)
            formula_credit = "ROUND(B2-C2,0)=0"  # Events = Journal = NetSuite
            summarySheet.conditional_formatting.add(
                f"B2:B{last_row}",
                FormulaRule(formula=[formula_credit], stopIfTrue=True, fill=green_fill)
            )
            summarySheet.conditional_formatting.add(
                f"B2:B{last_row}",
                FormulaRule(formula=[f"NOT({formula_credit})"], stopIfTrue=True, fill=red_fill)
            )
            summarySheet.conditional_formatting.add(
                f"C2:C{last_row}",
                FormulaRule(formula=[formula_credit], stopIfTrue=True, fill=green_fill)
            )
            summarySheet.conditional_formatting.add(
                f"C2:C{last_row}",
                FormulaRule(formula=[f"NOT({formula_credit})"], stopIfTrue=True, fill=red_fill)
            )

            # Conditional Formatting for Debit Columns (E, F, G)
            formula_debit = "ROUND(E2-F2,2)=0"  # Events = Journal = NetSuite
            summarySheet.conditional_formatting.add(
                f"E2:E{last_row}",
                FormulaRule(formula=[formula_debit], stopIfTrue=True, fill=green_fill)
            )
            summarySheet.conditional_formatting.add(
                f"E2:E{last_row}",
                FormulaRule(formula=[f"NOT({formula_debit})"], stopIfTrue=True, fill=red_fill)
            )
            summarySheet.conditional_formatting.add(
                f"F2:F{last_row}",
                FormulaRule(formula=[formula_debit], stopIfTrue=True, fill=green_fill)
            )
            summarySheet.conditional_formatting.add(
                f"F2:F{last_row}",
                FormulaRule(formula=[f"NOT({formula_debit})"], stopIfTrue=True, fill=red_fill)
            )
            
            return 'Event summary sheet created successfully'
        except Exception as e:
            print(traceback.format_exc())
            return f"Error creating Excel workbook: {e}"
         
    def createExecutionDetailsSheet(self, workbook: Workbook, bigQueryJobs: dict):
        '''Create execution details sheet for a journal'''
        print("Creating Execution Details sheet")
        try:
            # Retrieve job details
            jobDetails = self.GCPHandler.retrieveJobDetails(bigQueryJobs=bigQueryJobs)
            if not jobDetails:
                return "No job details found. Execution Details sheet not created."

            sheetName = "Execution Details"
            jobDetails = pd.DataFrame(jobDetails)
            tableNames = bigQueryJobs.values() 
            sheetNames = [self.getSheetName(tableName) for tableName in tableNames]

            jobDetails['Table Name'] = tableNames
            jobDetails['Sheet Name'] = sheetNames

            # Put tableNames and sheetNames in first two columns
            jobDetails = jobDetails[['Sheet Name','Table Name'] + [col for col in jobDetails.columns if col not in ['Table Name', 'Sheet Name']]]

            self.writeWorksheet(workbook, sheetName, pd.DataFrame(jobDetails))

            return "Execution details sheet created successfully"

        except Exception as e:
            print(traceback.format_exc())
            return f"Error creating Execution Details sheet: {e}"


    def createExcelWorkbook(self, journal: str, dateFrom: str, dateTo: str):
        '''Create Excel workbook for a journal'''
        try:
            workbook_path = f"temp/{journal}_{dateFrom}_{dateTo}.xlsx".replace("-", "_")

            # Ensure temp folder exists
            os.makedirs("temp", exist_ok=True)

            # Delete any existing files for this journal & date range
            for file in os.listdir("temp"):
                if file.startswith(f"{journal}_") and file.endswith(".xlsm"):
                    os.remove(os.path.join("temp", file))

            wb = Workbook()

            eventTables = [event['event_table'] for event in self.journalEventMapping[journal]['events']]
            eventsData = {}
            bigQueryJobs = {}

            for eventTable in eventTables:
                print(f"Getting data for {eventTable}")
                eventData = self.getTableData(tableName= eventTable, tableType= 'event', date_from= dateFrom, date_to= dateTo, bigQueryJobs= bigQueryJobs)
                self.writeWorksheet(wb, self.getSheetName(eventTable), eventData)
                eventsData[eventTable] = {'data': eventData.head(), 'length': eventData.shape[0]}  # Reduce RAM usage

            print(f"Getting data for {journal}")
            journalData = self.getTableData(journal, 'journal', dateFrom, dateTo, bigQueryJobs)
            self.writeWorksheet(wb, self.getSheetName(journal), journalData)
            journalData = {'data': journalData.head(), 'length': journalData.shape[0]}   # Reduce RAM usage

            accountCalculations = self.getAccountsCalculation(journal)
            accountCalculations.sort(key=lambda x: x['accountName'])

            self.createSummarySheet(wb, journal, journalData, eventsData, accountCalculations)
            self.createExecutionDetailsSheet(wb, bigQueryJobs)


            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Save and close workbook
            print(f"Saving workbook to {workbook_path}")
            time.sleep(1)  # Wait for workbook to be saved
            excelFile = BytesIO()
            wb.save(excelFile)
            excelFile.seek(0)

            return excelFile, workbook_path

        except Exception as e:
            traceback.print_exc()
            raise Exception(f"Error creating Excel workbook: {e}")      
